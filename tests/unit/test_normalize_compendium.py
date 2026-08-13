"""normalize_compendium Contract Test (ADR-0009/0010 + Amendment).

실제 FAL50 원본, Local Restricted Manifest, Actual Report·Authorization·
Binding에는 절대 접근하지 않는다. 모든 입력은 tmp_path의 Synthetic Fixture이며
TEST/FALTEST/urn:test 표기만 사용한다.
"""

import datetime
import hashlib
import inspect
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import openpyxl  # type: ignore[import-untyped]
import pytest
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import inspect_excel  # noqa: E402
import normalize_compendium  # noqa: E402
import validate_normalization_authorization as authz_module  # noqa: E402

WORKBOOK_NAME = "TEST-MARKER-WORKBOOK.xlsx"

DEFAULT_ROWS: list[list[Any]] = [
    ["TEST-HEADER-A", "TEST-HEADER-B", "TEST-HEADER-C"],
    ["TEST-ID-001", "TEST-NAME-001", 10],
    ["TEST-ID-002", "TEST-NAME-002", 20],
]

ARTIFACT_NAMES = (
    "normalized-records.local.json",
    "normalization-findings.local.json",
    "mapping-evidence.local.json",
    "normalization-summary.local.json",
)


def sha256_of_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def default_spec(**overrides: Any) -> dict[str, Any]:
    spec: dict[str, Any] = {
        "version": 1,
        "dataset_id": "TEST-DATASET-001",
        "source_sheet_ordinal": 0,
        "header_row": 1,
        "first_data_row": 2,
        "columns": [
            {
                "source_column_ordinal": 1,
                "target_field": "test_identifier",
                "required": True,
                "value_type": "string",
            },
            {
                "source_column_ordinal": 2,
                "target_field": "test_name",
                "required": True,
                "value_type": "string",
            },
            {
                "source_column_ordinal": 3,
                "target_field": "test_number",
                "required": False,
                "value_type": "number",
            },
        ],
    }
    spec.update(overrides)
    return spec


def build_authorization(report: dict[str, Any], report_bytes: bytes) -> dict[str, Any]:
    """Report 전체 Sheet Coverage와 Finding 정확 일치를 갖춘 Synthetic 승인."""
    sheets: list[dict[str, Any]] = []
    for sheet in report.get("sheets", []):
        confidence = sheet.get("headerConfidence")
        header_row = sheet.get("inferredHeaderRow")
        if confidence in ("high", "medium") and header_row is not None:
            sheets.append(
                {
                    "sheetOrdinal": sheet["sheetOrdinal"],
                    "classification": "data_table",
                    "normalize": True,
                    "headerRow": header_row,
                    "headerConfidence": confidence,
                    "mediumConfidenceApproved": confidence == "medium",
                    "exclusionReasonCode": None,
                }
            )
        else:
            sheets.append(
                {
                    "sheetOrdinal": sheet["sheetOrdinal"],
                    "classification": "excluded_non_data",
                    "normalize": False,
                    "headerRow": None,
                    "headerConfidence": confidence if confidence else "none",
                    "mediumConfidenceApproved": False,
                    "exclusionReasonCode": "TEST_EXCLUSION_001",
                }
            )
    acks = [
        {
            "code": finding["code"],
            "sheetOrdinal": authz_module._sheet_ordinal_from_path(finding.get("path")),
            "disposition": "accepted_for_reviewed_scope",
            "reasonCode": "TEST_REASON_001",
        }
        for finding in report.get("findings", [])
    ]
    return {
        "version": 1,
        "sourceId": report.get("sourceId"),
        "inspectionReportId": authz_module.compute_inspection_report_id(report_bytes),
        "outputStorageClass": "internal-restricted",
        "approvedOutputRootId": "TEST-RESTRICTED-ROOT-001",
        "sheets": sheets,
        "acknowledgedFindings": acks,
        "humanReviewCompleted": True,
    }


def write_bundle(
    env: dict[str, Path],
    *,
    binding_root: Path | None = None,
    auth_mutator: Any = None,
) -> None:
    report_bytes = env["report"].read_bytes()
    report = json.loads(report_bytes.decode("utf-8"))
    auth = build_authorization(report, report_bytes)
    if auth_mutator is not None:
        auth_mutator(auth)
    env["auth"].write_text(
        json.dumps(auth, ensure_ascii=False, sort_keys=True), encoding="utf-8"
    )
    root = binding_root if binding_root is not None else env["root"]
    binding = {
        "version": 1,
        "rootId": "TEST-RESTRICTED-ROOT-001",
        "storageClass": "internal-restricted",
        "rootPath": str(root),
    }
    env["binding"].write_text(
        json.dumps(binding, ensure_ascii=False, sort_keys=True), encoding="utf-8"
    )


def make_env(
    tmp_path: Path,
    rows: list[list[Any]] | None = None,
    spec: dict[str, Any] | None = None,
    extra_sheet: bool = False,
) -> dict[str, Path]:
    base = tmp_path / "sourcebase"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    wb = openpyxl.Workbook()
    # openpyxl 기본 <workbookProtection/> 요소가 Inspector에서 Protection으로
    # 탐지되어 normalizationReady=false가 되므로 Synthetic Fixture에서 제거한다.
    wb.security = None
    ws = wb.active
    ws.title = "TEST-SHEET-ALPHA"
    for row in rows if rows is not None else DEFAULT_ROWS:
        ws.append(row)
    if extra_sheet:
        beta = wb.create_sheet("TEST-SHEET-BETA")
        for row in DEFAULT_ROWS:
            beta.append(row)
    wb.save(workbook)

    manifest = base / "source-manifest.yaml"
    manifest.write_text(
        yaml.safe_dump(
            {
                "version": 1,
                "sources": [
                    {
                        "source_id": "TEST-SOURCE-001",
                        "fal_version": "FALTEST",
                        "ontology_version": "0.0.0-test",
                        "profile_version": "kr-profile-0.0.0-test",
                        "source_file": f"files/{WORKBOOK_NAME}",
                        "source_hash": sha256_of_file(workbook),
                        "resource_uri": "urn:test:source:001",
                    }
                ],
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )

    report = inspect_excel.inspect_workbook(
        workbook,
        manifest_path=manifest,
        options=inspect_excel.InspectionOptions(
            max_rows_per_sheet=5000, max_columns_per_sheet=200
        ),
    )
    report_path = tmp_path / "inspection-report.local.json"
    report_path.write_text(inspect_excel.serialize_report(report), encoding="utf-8")

    spec_path = tmp_path / "mapping-spec.local.yaml"
    spec_path.write_text(
        yaml.safe_dump(spec if spec is not None else default_spec(), allow_unicode=True),
        encoding="utf-8",
    )

    out = tmp_path / "out"
    out.mkdir()
    env = {
        "workbook": workbook,
        "manifest": manifest,
        "base": base,
        "report": report_path,
        "spec": spec_path,
        "out": out,
        "auth": tmp_path / "authorization.local.json",
        "binding": tmp_path / "output-root-binding.local.json",
        "root": tmp_path,
    }
    write_bundle(env)
    return env


def mutate_report(env: dict[str, Path], mutate: Any) -> None:
    report = json.loads(env["report"].read_text(encoding="utf-8"))
    mutate(report)
    env["report"].write_text(inspect_excel.serialize_report(report), encoding="utf-8")
    write_bundle(env)


def run_norm(env: dict[str, Path]) -> dict[str, Any]:
    return normalize_compendium.normalize_compendium(
        workbook_path=env["workbook"],
        manifest_path=env["manifest"],
        source_base_dir=env["base"],
        inspection_report_path=env["report"],
        authorization_path=env["auth"],
        output_root_binding_path=env["binding"],
        mapping_spec_path=env["spec"],
        output_dir=env["out"],
    )


def finding_codes(result: dict[str, Any]) -> list[str]:
    return [finding["code"] for finding in result["findings"]]


def fake_validator_result(**overrides: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        flag: True for flag in normalize_compendium.REQUIRED_AUTHORIZATION_FLAGS
    }
    base.update(
        {
            "classification": "internal-restricted",
            "authorizedSheetOrdinals": [0],
            "blockingFindingCount": 0,
            "reviewedFindingCount": 0,
            "findings": [],
        }
    )
    base.update(overrides)
    return base


def patch_workbook_boom(monkeypatch: pytest.MonkeyPatch) -> None:
    def _boom(**kwargs: Any) -> Any:
        raise AssertionError("Authorization 승인 전에 Workbook을 열면 안 된다")

    monkeypatch.setattr(
        normalize_compendium, "openpyxl", SimpleNamespace(load_workbook=_boom)
    )


# --- A. Authorization 필수 입력 ---


def test_valid_authorization_bundle_succeeds(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    result = run_norm(env)
    assert result["completed"] is True
    assert result["summary"]["normalizedRecordCount"] == 2
    for name in ARTIFACT_NAMES:
        assert (env["out"] / name).is_file()


def test_missing_authorization_file_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["auth"].unlink()
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_FILE_NOT_FOUND" in finding_codes(result)


def test_missing_binding_file_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["binding"].unlink()
    result = run_norm(env)
    assert result["completed"] is False
    assert "OUTPUT_ROOT_BINDING_FILE_NOT_FOUND" in finding_codes(result)


def test_broken_authorization_json_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["auth"].write_text("{ broken", encoding="utf-8")
    result = run_norm(env)
    assert "AUTHORIZATION_JSON_INVALID" in finding_codes(result)


def test_broken_binding_json_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["binding"].write_text("{ broken", encoding="utf-8")
    result = run_norm(env)
    assert "OUTPUT_ROOT_BINDING_JSON_INVALID" in finding_codes(result)


def test_authorization_root_not_object_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["auth"].write_text("[1, 2]", encoding="utf-8")
    assert "AUTHORIZATION_JSON_INVALID" in finding_codes(run_norm(env))


def test_binding_root_not_object_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["binding"].write_text('"TEST"', encoding="utf-8")
    assert "OUTPUT_ROOT_BINDING_JSON_INVALID" in finding_codes(run_norm(env))


# --- B. Validator Flag ---


@pytest.mark.parametrize(
    "flag",
    [
        "valid",
        "reportIdentityMatched",
        "sheetCoverageComplete",
        "findingCoverageComplete",
        "outputRootAuthorized",
        "humanReviewCompleted",
    ],
)
def test_false_validator_flag_rejected(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, flag: str
) -> None:
    env = make_env(tmp_path)
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(**{flag: False}),
    )
    patch_workbook_boom(monkeypatch)
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_VALIDATION_FAILED" in finding_codes(result)


def test_missing_validator_flag_rejected(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    fake = fake_validator_result()
    del fake["outputRootAuthorized"]
    monkeypatch.setattr(
        normalize_compendium, "validate_authorization", lambda **kwargs: fake
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_FLAG_MISSING" in finding_codes(result)


def test_validator_exception_not_propagated(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)

    def _raise(**kwargs: Any) -> dict[str, Any]:
        raise RuntimeError("TEST-VALIDATOR-CRASH")

    monkeypatch.setattr(normalize_compendium, "validate_authorization", _raise)
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_VALIDATION_FAILED" in finding_codes(result)
    assert "TEST-VALIDATOR-CRASH" not in json.dumps(result, ensure_ascii=False)


# --- C. File I/O 선행 차단 ---


def test_authorization_failure_blocks_all_io(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    # Identity 불일치를 유발한다.
    write_bundle(env, auth_mutator=lambda auth: auth.update(inspectionReportId="1" * 64))

    def _manifest_boom(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("Authorization 실패 시 source_manifest_load를 호출하면 안 된다")

    def _write_boom(*args: Any, **kwargs: Any) -> None:
        raise AssertionError("Authorization 실패 시 write_artifacts_atomic을 호출하면 안 된다")

    monkeypatch.setattr(normalize_compendium, "source_manifest_load", _manifest_boom)
    monkeypatch.setattr(normalize_compendium, "write_artifacts_atomic", _write_boom)
    patch_workbook_boom(monkeypatch)
    result = run_norm(env)
    assert result["completed"] is False
    assert list(env["out"].iterdir()) == []


def test_binding_failure_blocks_workbook_open(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    other_root = tmp_path / "other-root"
    other_root.mkdir()
    write_bundle(env, binding_root=other_root)  # out은 other_root 밖 → binding 실패
    patch_workbook_boom(monkeypatch)
    assert run_norm(env)["completed"] is False


def test_sheet_coverage_failure_blocks_workbook_open(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    write_bundle(env, auth_mutator=lambda auth: auth["sheets"].append(
        {
            "sheetOrdinal": 7,
            "classification": "excluded_non_data",
            "normalize": False,
            "headerRow": None,
            "headerConfidence": "none",
            "mediumConfidenceApproved": False,
            "exclusionReasonCode": "TEST_EXCLUSION_001",
        }
    ))
    patch_workbook_boom(monkeypatch)
    assert run_norm(env)["completed"] is False


def test_finding_coverage_failure_blocks_workbook_open(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    write_bundle(env, auth_mutator=lambda auth: auth["acknowledgedFindings"].append(
        {
            "code": "WORKBOOK_PROTECTION_ENABLED",
            "sheetOrdinal": None,
            "disposition": "accepted_for_reviewed_scope",
            "reasonCode": "TEST_REASON_001",
        }
    ))
    patch_workbook_boom(monkeypatch)
    assert run_norm(env)["completed"] is False


# --- D. Mapping Authorization ---


def test_authorized_mapping_sheet_succeeds(tmp_path: Path) -> None:
    assert run_norm(make_env(tmp_path))["completed"] is True


def test_unauthorized_mapping_sheet_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path, extra_sheet=True, spec=default_spec(source_sheet_ordinal=1))
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["sheets"][1].update(
            classification="excluded_non_data",
            normalize=False,
            headerRow=None,
            headerConfidence="none",
            exclusionReasonCode="TEST_EXCLUSION_001",
        ),
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "MAPPING_SHEET_NOT_AUTHORIZED" in finding_codes(result)


def test_mapping_header_row_mismatch_with_authorization_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path, spec=default_spec(header_row=3, first_data_row=4))
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_HEADER_MISMATCH" in finding_codes(result)


def test_excluded_sheet_mapping_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(
        env,
        lambda r: r["sheets"][0].update(inferredHeaderRow=None, headerConfidence="none"),
    )
    assert run_norm(env)["completed"] is False


def test_code_list_sheet_mapping_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["sheets"][0].update(
            classification="code_list",
            normalize=False,
            headerRow=None,
            headerConfidence="none",
            exclusionReasonCode=None,
        ),
    )
    assert run_norm(env)["completed"] is False


def test_unsorted_authorized_ordinals_still_work(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(authorizedSheetOrdinals=[2, 0, 1]),
    )
    assert run_norm(env)["completed"] is True


# --- E. Medium Confidence 우회 제거 ---


def test_library_api_has_no_medium_override() -> None:
    assert not hasattr(normalize_compendium, "NormalizationOptions")
    params = inspect.signature(normalize_compendium.normalize_compendium).parameters
    assert "authorization_path" in params
    assert "output_root_binding_path" in params
    assert "options" not in params


def test_cli_medium_override_is_removed(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    argv = [
        str(env["workbook"]),
        "--manifest", str(env["manifest"]),
        "--source-base-dir", str(env["base"]),
        "--inspection-report", str(env["report"]),
        "--authorization", str(env["auth"]),
        "--output-root-binding", str(env["binding"]),
        "--mapping-spec", str(env["spec"]),
        "--output-dir", str(env["out"]),
        "--allow-medium-header-confidence",
    ]
    with pytest.raises(SystemExit) as excinfo:
        normalize_compendium.main(argv)
    assert excinfo.value.code == 2


def test_approved_medium_confidence_succeeds(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["sheets"][0].update(headerConfidence="medium"))
    assert run_norm(env)["completed"] is True


def test_unapproved_medium_confidence_fails_at_validator(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["sheets"][0].update(headerConfidence="medium"))
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["sheets"][0].update(
            mediumConfidenceApproved=False
        ),
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "AUTHORIZATION_VALIDATION_FAILED" in finding_codes(result)


def test_normalizer_has_no_medium_inference() -> None:
    source = Path(normalize_compendium.__file__).read_text(encoding="utf-8")
    assert "allow_medium_header_confidence" not in source


# --- F. Repository Root 및 Output ---


def test_repository_root_detection() -> None:
    assert normalize_compendium._detect_repository_root() == REPO_ROOT


def test_missing_git_root_fails(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    env = make_env(tmp_path)
    monkeypatch.setattr(
        normalize_compendium, "_repository_root_candidate", lambda: tmp_path / "no-repo"
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "REPOSITORY_ROOT_NOT_FOUND" in finding_codes(result)


def test_missing_pyproject_fails(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    env = make_env(tmp_path)
    fake_root = tmp_path / "fake-repo"
    (fake_root / ".git").mkdir(parents=True)
    monkeypatch.setattr(
        normalize_compendium, "_repository_root_candidate", lambda: fake_root
    )
    result = run_norm(env)
    assert "REPOSITORY_ROOT_NOT_FOUND" in finding_codes(result)


def test_repo_normalized_ignored_output_allowed(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    out = REPO_ROOT / "data" / "normalized" / f"TEST-NORM-{os.getpid()}"
    out.mkdir(parents=True)
    try:
        env["out"] = out
        write_bundle(env, binding_root=REPO_ROOT / "data" / "normalized")
        result = run_norm(env)
        assert result["completed"] is True
        for name in ARTIFACT_NAMES:
            rel = (out / name).relative_to(REPO_ROOT).as_posix()
            check = subprocess.run(
                ["git", "check-ignore", rel], capture_output=True, cwd=str(REPO_ROOT)
            )
            assert check.returncode == 0
    finally:
        shutil.rmtree(out, ignore_errors=True)


def test_defense_in_depth_conflict_still_fails(tmp_path: Path) -> None:
    # Binding Root 안이지만 Workbook Directory로의 출력은 Normalizer Boundary가 막는다.
    env = make_env(tmp_path)
    env["out"] = env["base"] / "files"
    result = run_norm(env)
    assert result["completed"] is False
    assert "OUTPUT_PATH_CONFLICT" in finding_codes(result)


def test_repo_root_output_rejected(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["out"] = REPO_ROOT
    assert run_norm(env)["completed"] is False


def test_symlink_escape_rejected(tmp_path: Path) -> None:
    outside = tmp_path / "escape-target"
    outside.mkdir()
    link = REPO_ROOT / "data" / "normalized" / f"TEST-LINK-{os.getpid()}"
    try:
        os.symlink(outside, link, target_is_directory=True)
    except (OSError, NotImplementedError):
        pytest.skip("이 환경에서는 Symlink 생성이 지원되지 않는다")
    try:
        env = make_env(tmp_path)
        env["out"] = link
        write_bundle(env, binding_root=REPO_ROOT / "data" / "normalized")
        assert run_norm(env)["completed"] is False
    finally:
        link.unlink(missing_ok=True)


def test_repository_root_not_exposed(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    env["out"] = REPO_ROOT
    serialized = json.dumps(run_norm(env), ensure_ascii=False, sort_keys=True)
    assert str(REPO_ROOT).lower() not in serialized.lower()


def mutate_report_only(env: dict[str, Path], mutate: Any) -> None:
    """Bundle 재생성 없이 Report만 변경한다 (fake Validator와 함께 Compatibility 검사용)."""
    report = json.loads(env["report"].read_text(encoding="utf-8"))
    mutate(report)
    env["report"].write_text(inspect_excel.serialize_report(report), encoding="utf-8")


def add_report_finding(code: str, path: str = "$.workbook") -> Any:
    def _mutate(report: dict[str, Any]) -> None:
        report["findings"].append(
            {
                "severity": "WARNING",
                "code": code,
                "message": "TEST",
                "path": path,
                "actualValue": None,
            }
        )
        report["summary"]["normalizationReady"] = False

    return _mutate


# --- H. Authorization-aware Readiness ---


def test_reviewable_finding_with_ready_false_succeeds(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, add_report_finding("WORKBOOK_DECLARED_DIMENSION_EXCESSIVE"))
    result = run_norm(env)
    assert result["completed"] is True
    # 실제 Inspection Finding Message·Count는 Artifact에 복사되지 않는다.
    for name in ARTIFACT_NAMES:
        text = (env["out"] / name).read_text(encoding="utf-8")
        assert "WORKBOOK_DECLARED_DIMENSION_EXCESSIVE" not in text


def test_reviewable_finding_with_ready_true_succeeds(tmp_path: Path) -> None:
    env = make_env(tmp_path)

    def _mutate(report: dict[str, Any]) -> None:
        add_report_finding("WORKBOOK_DECLARED_DIMENSION_EXCESSIVE")(report)
        report["summary"]["normalizationReady"] = True

    mutate_report(env, _mutate)
    assert run_norm(env)["completed"] is True


def test_reviewable_remains_blocking_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, add_report_finding("WORKBOOK_DECLARED_DIMENSION_EXCESSIVE"))
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["acknowledgedFindings"][0].update(
            disposition="remains_blocking"
        ),
    )
    assert run_norm(env)["completed"] is False


def test_reviewable_missing_authorization_fails(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, add_report_finding("WORKBOOK_DECLARED_DIMENSION_EXCESSIVE"))
    write_bundle(env, auth_mutator=lambda auth: auth.update(acknowledgedFindings=[]))
    assert run_norm(env)["completed"] is False


@pytest.mark.parametrize(
    "disposition", ["resolved", "accepted_for_reviewed_scope", "remains_blocking"]
)
def test_blocking_finding_fails_for_any_disposition(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, disposition: str
) -> None:
    env = make_env(tmp_path)
    mutate_report(env, add_report_finding("WORKBOOK_SCAN_LIMIT_REACHED"))
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["acknowledgedFindings"][0].update(
            disposition=disposition
        ),
    )

    def _manifest_boom(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("Blocking 실패 시 source_manifest_load를 호출하면 안 된다")

    def _write_boom(*args: Any, **kwargs: Any) -> None:
        raise AssertionError("Blocking 실패 시 write_artifacts_atomic을 호출하면 안 된다")

    monkeypatch.setattr(normalize_compendium, "source_manifest_load", _manifest_boom)
    monkeypatch.setattr(normalize_compendium, "write_artifacts_atomic", _write_boom)
    patch_workbook_boom(monkeypatch)
    result = run_norm(env)
    assert result["completed"] is False
    assert list(env["out"].iterdir()) == []


def test_ready_false_alone_does_not_block(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["summary"].update(normalizationReady=False))
    assert run_norm(env)["completed"] is True


def test_ready_true_does_not_override_validator(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["summary"].update(normalizationReady=True))
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(valid=False),
    )
    assert run_norm(env)["completed"] is False


def test_missing_ready_field_is_not_a_failure(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["summary"].pop("normalizationReady", None))
    assert run_norm(env)["completed"] is True


def test_wrong_type_ready_field_is_ignored(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    mutate_report(env, lambda r: r["summary"].update(normalizationReady="TEST"))
    assert run_norm(env)["completed"] is True


# --- I. Sheet Ordinal Compatibility ---


def test_non_contiguous_ordinals_succeed(tmp_path: Path) -> None:
    env = make_env(tmp_path, extra_sheet=True)
    mutate_report(env, lambda r: r["sheets"][1].update(sheetOrdinal=5))
    assert run_norm(env)["completed"] is True


def test_shuffled_sheet_list_order_succeeds(tmp_path: Path) -> None:
    env = make_env(tmp_path, extra_sheet=True)

    def _mutate(report: dict[str, Any]) -> None:
        report["sheets"][1]["sheetOrdinal"] = 5
        report["sheets"].reverse()

    mutate_report(env, _mutate)
    assert run_norm(env)["completed"] is True


def test_ordinal_is_not_treated_as_list_index(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path, extra_sheet=True)

    def _mutate(report: dict[str, Any]) -> None:
        # List 첫 항목은 ordinal 5·headerRow 3, ordinal 0은 List 뒤쪽에 있다.
        report["sheets"][0]["sheetOrdinal"] = 5
        report["sheets"][0]["inferredHeaderRow"] = 3
        report["sheets"][1]["sheetOrdinal"] = 0

    mutate_report_only(env, _mutate)
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(authorizedSheetOrdinals=[0, 5]),
    )
    # ordinal 0의 headerRow는 1 — Index로 조회하면 ordinal 5(row 3)를 잘못 얻는다.
    assert run_norm(env)["completed"] is True


def test_mapping_ordinal_missing_in_report_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path, spec=default_spec(source_sheet_ordinal=7))
    write_bundle(
        env,
        auth_mutator=lambda auth: auth["sheets"].append(
            {
                "sheetOrdinal": 7,
                "classification": "data_table",
                "normalize": True,
                "headerRow": 1,
                "headerConfidence": "high",
                "mediumConfidenceApproved": False,
                "exclusionReasonCode": None,
            }
        ),
    )
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(authorizedSheetOrdinals=[0, 7]),
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "INSPECTION_SHEET_OUT_OF_RANGE" in finding_codes(result)


def _compat_env(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, mutate: Any
) -> dict[str, Path]:
    env = make_env(tmp_path)
    mutate_report_only(env, mutate)
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(),
    )
    return env


def test_report_duplicate_ordinal_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = _compat_env(
        tmp_path,
        monkeypatch,
        lambda r: r["sheets"].append(dict(r["sheets"][0])),
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "INSPECTION_SHEET_ORDINAL_DUPLICATE" in finding_codes(result)


def test_report_bool_ordinal_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = _compat_env(
        tmp_path, monkeypatch, lambda r: r["sheets"][0].update(sheetOrdinal=True)
    )
    result = run_norm(env)
    assert "INSPECTION_SHEET_ORDINAL_INVALID" in finding_codes(result)


def test_report_string_ordinal_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = _compat_env(
        tmp_path, monkeypatch, lambda r: r["sheets"][0].update(sheetOrdinal="TEST")
    )
    assert "INSPECTION_SHEET_ORDINAL_INVALID" in finding_codes(run_norm(env))


def test_report_negative_ordinal_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = _compat_env(
        tmp_path, monkeypatch, lambda r: r["sheets"][0].update(sheetOrdinal=-1)
    )
    assert "INSPECTION_SHEET_ORDINAL_INVALID" in finding_codes(run_norm(env))


def test_report_non_object_sheet_item_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = _compat_env(
        tmp_path, monkeypatch, lambda r: r["sheets"].append("TEST-NOT-A-SHEET")
    )
    result = run_norm(env)
    assert result["completed"] is False
    assert "INSPECTION_REPORT_INVALID" in finding_codes(result)


def test_compat_header_mismatch_fails_and_blocks_workbook(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path, spec=default_spec(header_row=3, first_data_row=4))
    write_bundle(
        env, auth_mutator=lambda auth: auth["sheets"][0].update(headerRow=3)
    )
    monkeypatch.setattr(
        normalize_compendium,
        "validate_authorization",
        lambda **kwargs: fake_validator_result(),
    )
    patch_workbook_boom(monkeypatch)
    result = run_norm(env)
    assert result["completed"] is False
    assert "INSPECTION_HEADER_MISMATCH" in finding_codes(result)


# --- J. Console Count 비노출 ---


FORBIDDEN_CONSOLE_TOKENS = (
    "normalizedRecordCount",
    "rejectedRecordCount",
    "findingCount",
    "sheetOrdinal",
    "TEST-DATASET",
    "TEST-SOURCE",
    WORKBOOK_NAME,
)


def cli_argv(env: dict[str, Path]) -> list[str]:
    return [
        str(env["workbook"]),
        "--manifest", str(env["manifest"]),
        "--source-base-dir", str(env["base"]),
        "--inspection-report", str(env["report"]),
        "--authorization", str(env["auth"]),
        "--output-root-binding", str(env["binding"]),
        "--mapping-spec", str(env["spec"]),
        "--output-dir", str(env["out"]),
    ]


def test_success_console_contract(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    env = make_env(tmp_path)
    assert normalize_compendium.main(cli_argv(env)) == 0
    out = capsys.readouterr().out
    for token in FORBIDDEN_CONSOLE_TOKENS:
        assert token not in out
    assert "artifactSetCreated=True" in out
    assert "humanReviewRequired=False" in out
    assert str(tmp_path).lower() not in out.lower()


def test_failure_console_contract(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    env = make_env(tmp_path)
    env["auth"].unlink()
    assert normalize_compendium.main(cli_argv(env)) == 1
    out = capsys.readouterr().out
    for token in FORBIDDEN_CONSOLE_TOKENS:
        assert token not in out
    assert "artifactSetCreated=False" in out
    assert "humanReviewRequired=True" in out


def test_rejected_rows_console_review_flag(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    env = type_env(tmp_path, None, "string", required=True)
    assert normalize_compendium.main(cli_argv(env)) == 0
    out = capsys.readouterr().out
    assert "humanReviewRequired=True" in out
    assert "rejectedRecordCount" not in out


# --- G. Regression: Type·Row·Atomic·결정론 ---


def type_env(
    tmp_path: Path, value: Any, value_type: str, required: bool = False
) -> dict[str, Path]:
    rows: list[list[Any]] = [["TEST-HEADER-A", "TEST-HEADER-B"], ["TEST-ID-001", value]]
    spec = default_spec()
    spec["columns"] = [
        {
            "source_column_ordinal": 1,
            "target_field": "test_identifier",
            "required": True,
            "value_type": "string",
        },
        {
            "source_column_ordinal": 2,
            "target_field": "test_value",
            "required": required,
            "value_type": value_type,
        },
    ]
    return make_env(tmp_path, rows=rows, spec=spec)


def first_value(result: dict[str, Any]) -> Any:
    records = result["artifacts"]["normalized-records.local.json"]["records"]
    return records[0]["values"]["test_value"]


def test_type_normalization_regression(tmp_path: Path) -> None:
    assert first_value(run_norm(type_env(tmp_path / "s", "  TEST-TRIM  ", "string"))) == "TEST-TRIM"
    assert first_value(run_norm(type_env(tmp_path / "n", 12.5, "number"))) == 12.5
    assert first_value(run_norm(type_env(tmp_path / "i", 10, "integer"))) == 10
    assert first_value(run_norm(type_env(tmp_path / "b", True, "boolean"))) is True
    assert (
        first_value(run_norm(type_env(tmp_path / "d", datetime.date(2020, 1, 2), "date")))
        == "2020-01-02"
    )


def test_rejection_regression(tmp_path: Path) -> None:
    formula_env = type_env(tmp_path / "f", "=SUM(1,2)", "number")
    mutate_report(
        formula_env,
        lambda r: (r.update(findings=[]), r["summary"].update(normalizationReady=True)),
    )
    result = run_norm(formula_env)
    assert result["summary"]["rejectedRecordCount"] == 1
    assert "FORMULA_VALUE_PROHIBITED" in finding_codes(result)

    missing = run_norm(type_env(tmp_path / "m", None, "string", required=True))
    assert missing["summary"]["rejectedRecordCount"] == 1
    assert "REQUIRED_VALUE_MISSING" in finding_codes(missing)


def test_existing_output_rejected(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    (env["out"] / ARTIFACT_NAMES[0]).write_text("{}\n", encoding="utf-8")
    result = run_norm(env)
    assert result["completed"] is False
    assert "OUTPUT_ALREADY_EXISTS" in finding_codes(result)


def test_partial_write_failure_cleans_up(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env = make_env(tmp_path)
    original = normalize_compendium.serialize_artifact
    calls = {"count": 0}

    def _fail_on_third(data: dict[str, object]) -> str:
        calls["count"] += 1
        if calls["count"] >= 3:
            raise ValueError("TEST-SERIALIZE-FAILURE")
        return original(data)

    monkeypatch.setattr(normalize_compendium, "serialize_artifact", _fail_on_third)
    result = run_norm(env)
    assert result["completed"] is False
    assert "OUTPUT_WRITE_FAILED" in finding_codes(result)
    assert list(env["out"].iterdir()) == []


def test_inputs_unchanged(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    keys = ("workbook", "manifest", "report", "spec", "auth", "binding")
    before = {key: sha256_of_file(env[key]) for key in keys}
    run_norm(env)
    after = {key: sha256_of_file(env[key]) for key in keys}
    assert before == after


def test_same_input_same_artifact_bytes(tmp_path: Path) -> None:
    env_a = make_env(tmp_path / "a")
    env_b = make_env(tmp_path / "b")
    result_a = run_norm(env_a)
    result_b = run_norm(env_b)
    assert result_a["summary"] == result_b["summary"]
    for name in ARTIFACT_NAMES:
        assert (env_a["out"] / name).read_bytes() == (env_b["out"] / name).read_bytes()


def test_artifacts_have_no_disallowed_content(tmp_path: Path) -> None:
    env = make_env(tmp_path)
    run_norm(env)
    workbook_hash = sha256_of_file(env["workbook"])
    report_digest = authz_module.compute_inspection_report_id(env["report"].read_bytes())
    for name in ARTIFACT_NAMES:
        text = (env["out"] / name).read_text(encoding="utf-8")
        assert "generatedAt" not in text and "timestamp" not in text
        assert str(tmp_path).lower() not in text.lower()
        assert "c:\\" not in text.lower() and "c:/" not in text.lower()
        assert workbook_hash not in text
        assert report_digest not in text
        assert WORKBOOK_NAME not in text
        artifact = json.loads(text)
        assert artifact["classification"] == "internal-restricted"


def test_failure_findings_and_console_contract(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    env = make_env(tmp_path)
    write_bundle(env, auth_mutator=lambda auth: auth.update(inspectionReportId="1" * 64))
    failure = run_norm(env)
    for finding in failure["findings"]:
        assert finding["actualValue"] is None

    success_env = make_env(tmp_path / "cli")
    cli_out = tmp_path / "cli" / "cli-out"
    cli_out.mkdir()
    success_env["out"] = cli_out
    argv = [
        str(success_env["workbook"]),
        "--manifest", str(success_env["manifest"]),
        "--source-base-dir", str(success_env["base"]),
        "--inspection-report", str(success_env["report"]),
        "--authorization", str(success_env["auth"]),
        "--output-root-binding", str(success_env["binding"]),
        "--mapping-spec", str(success_env["spec"]),
        "--output-dir", str(cli_out),
    ]
    assert normalize_compendium.main(argv) == 0
    out = capsys.readouterr().out
    assert "TEST-ID-001" not in out
    assert "TEST-HEADER" not in out
    assert WORKBOOK_NAME not in out
    assert str(tmp_path).lower() not in out.lower()
    assert "internal-restricted" in out
