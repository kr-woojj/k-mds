"""inspect_excel Contract Test (ADR-0007).

실제 FAL50 원본에는 절대 접근하지 않는다. 모든 Workbook은 tmp_path에서
openpyxl과 zipfile로 생성한 Synthetic Fixture이며 TEST/FALTEST/urn:test
표기만 사용한다.
"""

import datetime
import hashlib
import sys
import zipfile
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
import pytest
import yaml
from openpyxl.comments import Comment  # type: ignore[import-untyped]
from openpyxl.styles import Protection  # type: ignore[import-untyped]  # noqa: F401
from openpyxl.workbook.protection import WorkbookProtection  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table  # type: ignore[import-untyped]

REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = REPO_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import inspect_excel  # noqa: E402

SHEET_NAME = "TEST-SHEET-ALPHA"
WORKBOOK_NAME = "TEST-MARKER-WORKBOOK.xlsx"

FORBIDDEN_REPORT_KEYS = {
    "inputPath", "outputPath", "filename", "fileName", "sourceFile", "sourceHash",
    "resourceUri", "rawValues", "cellValues", "sheetName", "headerText", "formulaText",
    "commentText", "hyperlinkTarget", "externalLinkTarget", "namedRangeFormula",
    "creatorValue", "lastModifiedByValue", "companyValue", "timestamp", "generatedAt",
    "username", "machineName", "workingDirectory", "absolutePath",
}


def build_normal_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "TEST-HEADER-A"
    ws["B1"] = "TEST-HEADER-B"
    ws["C1"] = "TEST-HEADER-C"
    ws["A2"] = "TEST-CELL-VALUE"
    ws["B2"] = 42
    ws["C2"] = True
    ws["D2"] = datetime.date(2020, 1, 1)
    ws["A3"] = "TEST-CELL-2"
    ws["B3"] = 3.14
    wb.save(path)


def build_structural_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "TEST-HEADER-A"
    ws["B1"] = "TEST-HEADER-B"
    ws["A2"] = "TEST-CELL-VALUE"
    ws.merge_cells("A10:B10")
    ws.row_dimensions[5].hidden = True
    ws.column_dimensions["D"].hidden = True
    ws["A3"].comment = Comment("TEST-COMMENT-TEXT", "TEST-AUTHOR")
    ws["A4"] = "TEST-LINK"
    ws["A4"].hyperlink = "urn:test:hyperlink-target"
    validation = DataValidation(type="list", formula1='"TESTX,TESTY"')
    ws.add_data_validation(validation)
    validation.add("B4")
    ws["E1"] = "TEST-T1"
    ws["F1"] = "TEST-T2"
    ws["E2"] = 1
    ws["F2"] = 2
    ws.add_table(Table(displayName="TESTTABLE1", ref="E1:F2"))
    ws.protection.sheet = True
    hidden = wb.create_sheet("TEST-SHEET-HIDDEN")
    hidden["A1"] = "TEST-HIDDEN-CELL"
    hidden.sheet_state = "hidden"
    very_hidden = wb.create_sheet("TEST-SHEET-VERYHIDDEN")
    very_hidden["A1"] = "TEST-VERYHIDDEN-CELL"
    very_hidden.sheet_state = "veryHidden"
    wb.security = WorkbookProtection(lockStructure=True)
    wb.save(path)


def build_formula_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "TEST-HEADER-A"
    ws["B1"] = "TEST-HEADER-B"
    ws["A2"] = 1
    ws["B2"] = "=SUM(1,2)"
    wb.save(path)


def build_sparse_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = 7
    ws["C3"] = "TEST-LONE"
    ws["B5"] = 9
    wb.save(path)


def build_duplicate_header_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "TEST-HEADER-DUP"
    ws["B1"] = "TEST-HEADER-DUP"
    ws["C1"] = "TEST-HEADER-DUP"
    ws["A2"] = 1
    ws["B2"] = 2
    ws["C2"] = 3
    wb.save(path)


def build_merged_header_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "TEST-HEADER-MERGED"
    ws["C1"] = "TEST-HEADER-C"
    ws.merge_cells("A1:B1")
    ws["A2"] = 1
    ws["C2"] = 2
    wb.save(path)


def add_zip_entry(source: Path, target: Path, name: str, data: bytes) -> None:
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(target, "w") as zout:
        for info in zin.infolist():
            zout.writestr(info, zin.read(info.filename))
        zout.writestr(name, data)


def sha256_of_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_valid_manifest(base: Path, workbook: Path) -> Path:
    manifest_path = base / "source-manifest.yaml"
    manifest_path.write_text(
        yaml.safe_dump(
            {
                "version": 1,
                "sources": [
                    {
                        "source_id": "TEST-SOURCE-001",
                        "fal_version": "FALTEST",
                        "ontology_version": "0.0.0-test",
                        "profile_version": "kr-profile-0.0.0-test",
                        "source_file": workbook.relative_to(base).as_posix(),
                        "source_hash": sha256_of_file(workbook),
                        "resource_uri": "urn:test:source:001",
                    }
                ],
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    return manifest_path


def write_mismatch_manifest(base: Path, workbook: Path) -> Path:
    manifest_path = write_valid_manifest(base, workbook)
    text = manifest_path.read_text(encoding="utf-8")
    manifest_path.write_text(
        text.replace(sha256_of_file(workbook), "f" * 64), encoding="utf-8"
    )
    return manifest_path


def write_pending_manifest(base: Path) -> Path:
    manifest_path = base / "source-manifest.yaml"
    manifest_path.write_text(
        "standard:\n  fal_version: FALTEST\n  status: pending_source\n"
        "files: []\ningestion:\n  status: pending_source\n",
        encoding="utf-8",
    )
    return manifest_path


def make_fixture(tmp_path: Path, builder: Any = build_normal_workbook) -> tuple[Path, Path]:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    builder(workbook)
    manifest = write_valid_manifest(base, workbook)
    return workbook, manifest


def make_options(**overrides: Any) -> "inspect_excel.InspectionOptions":
    base: dict[str, Any] = {"max_rows_per_sheet": 50, "max_columns_per_sheet": 20}
    base.update(overrides)
    return inspect_excel.InspectionOptions(**base)


def run_inspect(workbook: Path, manifest: Path, **overrides: Any) -> dict[str, Any]:
    return inspect_excel.inspect_workbook(
        workbook, manifest_path=manifest, options=make_options(**overrides)
    )


def walk_keys(node: object) -> set[str]:
    keys: set[str] = set()
    if isinstance(node, dict):
        for key, value in node.items():
            keys.add(str(key))
            keys |= walk_keys(value)
    elif isinstance(node, list):
        for item in node:
            keys |= walk_keys(item)
    return keys


def finding_codes(report: dict[str, Any]) -> list[str]:
    return [finding["code"] for finding in report["findings"]]


# --- A. 원본 보호 및 결정론 ---


def test_normal_inspection_succeeds(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is True
    assert report["manifestStatus"] == "verified"


def test_workbook_sha256_unchanged(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    before = sha256_of_file(workbook)
    run_inspect(workbook, manifest)
    assert sha256_of_file(workbook) == before


def test_report_dict_deterministic(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    assert run_inspect(workbook, manifest) == run_inspect(workbook, manifest)


def test_serialize_bytes_deterministic(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    first = inspect_excel.serialize_report(run_inspect(workbook, manifest))
    second = inspect_excel.serialize_report(run_inspect(workbook, manifest))
    assert first.encode("utf-8") == second.encode("utf-8")
    assert first.endswith("\n")
    parsed = __import__("json").loads(first)
    assert isinstance(parsed, dict)


def test_report_has_no_timestamp_or_absolute_path(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    serialized = inspect_excel.serialize_report(run_inspect(workbook, manifest))
    assert "generatedAt" not in serialized and "timestamp" not in serialized
    assert str(tmp_path).lower() not in serialized.lower()
    assert "c:\\" not in serialized.lower() and "c:/" not in serialized.lower()


def test_report_has_no_raw_content(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_formula_workbook)
    serialized = inspect_excel.serialize_report(run_inspect(workbook, manifest))
    assert WORKBOOK_NAME not in serialized
    assert SHEET_NAME not in serialized
    assert "TEST-HEADER-A" not in serialized
    assert "TEST-CELL-VALUE" not in serialized
    assert "=SUM" not in serialized


def test_forbidden_keys_absent(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    report = run_inspect(workbook, manifest)
    assert walk_keys(report) & FORBIDDEN_REPORT_KEYS == set()


# --- B. Inventory ---


def test_sheet_counts(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    workbook_info = run_inspect(workbook, manifest)["workbook"]
    assert workbook_info["sheetCount"] == 3
    assert workbook_info["visibleSheetCount"] == 1
    assert workbook_info["hiddenSheetCount"] == 1
    assert workbook_info["veryHiddenSheetCount"] == 1


def test_sheet_order_and_ordinals(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    sheets = run_inspect(workbook, manifest)["sheets"]
    assert [sheet["sheetOrdinal"] for sheet in sheets] == [0, 1, 2]
    assert [sheet["state"] for sheet in sheets] == ["visible", "hidden", "veryHidden"]


def test_sheet_name_digest_is_deterministic(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    sheet = run_inspect(workbook, manifest)["sheets"][0]
    expected = hashlib.sha256(SHEET_NAME.encode("utf-8")).hexdigest()
    assert sheet["sheetNameDigest"] == expected


def test_non_empty_and_cell_type_counts(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    sheet = run_inspect(workbook, manifest)["sheets"][0]
    assert sheet["nonEmptyCellCount"] == 9
    counts = sheet["cellTypeCounts"]
    assert counts["string"] == 5
    assert counts["number"] == 2
    assert counts["boolean"] == 1
    assert counts["date"] == 1
    assert counts["blank"] == 3
    assert counts["formula"] == 0


def test_formula_count(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_formula_workbook)
    report = run_inspect(workbook, manifest)
    assert report["sheets"][0]["formulaCellCount"] == 1
    assert report["workbook"]["formulaCellCount"] == 1


def test_merged_range_count(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    assert run_inspect(workbook, manifest)["sheets"][0]["mergedRangeCount"] == 1


def test_hidden_row_and_column_counts(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    sheet = run_inspect(workbook, manifest)["sheets"][0]
    assert sheet["hiddenRowCount"] == 1
    assert sheet["hiddenColumnCount"] == 1


def test_comment_count(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    report = run_inspect(workbook, manifest)
    assert report["workbook"]["commentCount"] == 1
    assert report["sheets"][0]["commentCount"] == 1


def test_hyperlink_count_and_target_not_exposed(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    report = run_inspect(workbook, manifest)
    assert report["sheets"][0]["hyperlinkCount"] == 1
    assert "urn:test:hyperlink-target" not in inspect_excel.serialize_report(report)


def test_data_validation_count(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    assert run_inspect(workbook, manifest)["sheets"][0]["dataValidationCount"] == 1


def test_table_count(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    assert run_inspect(workbook, manifest)["sheets"][0]["tableCount"] == 1


def test_protection_detected(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_structural_workbook)
    report = run_inspect(workbook, manifest)
    assert report["workbook"]["hasSheetProtection"] is True
    assert report["workbook"]["hasWorkbookProtection"] is True
    assert "WORKBOOK_PROTECTION_ENABLED" in finding_codes(report)


# --- C. Header ---


def test_simple_header_row_detected(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    sheet = run_inspect(workbook, manifest)["sheets"][0]
    assert sheet["inferredHeaderRow"] == 1
    assert sheet["headerConfidence"] in ("high", "medium")


def test_header_candidates_are_digest_only(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    candidates = run_inspect(workbook, manifest)["sheets"][0]["headerCandidates"]
    assert len(candidates) == 3
    for candidate in candidates:
        assert set(candidate) == {"row", "column", "valueDigest", "normalizedLength", "cellType"}
        assert len(candidate["valueDigest"]) == 64
        assert "TEST-HEADER" not in str(candidate)


def test_duplicate_header_digest_detected(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_duplicate_header_workbook)
    report = run_inspect(workbook, manifest)
    assert report["sheets"][0]["duplicateHeaderDigestCount"] > 0
    assert "WORKBOOK_DUPLICATE_HEADER_DIGEST" in finding_codes(report)


def test_header_not_detected_finding(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_sparse_workbook)
    report = run_inspect(workbook, manifest)
    assert report["sheets"][0]["inferredHeaderRow"] is None
    assert report["sheets"][0]["headerConfidence"] == "none"
    assert "WORKBOOK_HEADER_NOT_DETECTED" in finding_codes(report)


def test_merged_header_finding(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_merged_header_workbook)
    assert "WORKBOOK_MERGED_HEADER_REGION" in finding_codes(run_inspect(workbook, manifest))


def test_header_confidence_is_valid_category(tmp_path: Path) -> None:
    for builder in (build_normal_workbook, build_sparse_workbook, build_duplicate_header_workbook):
        workbook, manifest = make_fixture(tmp_path / builder.__name__, builder)
        sheet = run_inspect(workbook, manifest)["sheets"][0]
        assert sheet["headerConfidence"] in ("high", "medium", "low", "none")


# --- D. ZIP 및 위험 Feature ---


def test_formula_is_warning_by_default(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_formula_workbook)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is True
    formula_findings = [
        finding for finding in report["findings"] if finding["code"] == "WORKBOOK_FORMULA_PRESENT"
    ]
    assert formula_findings and formula_findings[0]["severity"] == "WARNING"


def test_fail_on_formulas_cli_exit(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path, build_formula_workbook)
    output = tmp_path / "report.local.json"
    argv = [
        str(workbook), "--manifest", str(manifest),
        "--max-rows-per-sheet", "50", "--max-columns-per-sheet", "20",
        "--output", str(output), "--fail-on-formulas",
    ]
    assert inspect_excel.main(argv) == 1


def test_invalid_zip_fails(tmp_path: Path) -> None:
    base = tmp_path / "base"
    base.mkdir()
    workbook = base / WORKBOOK_NAME
    workbook.write_bytes(b"TEST-NOT-A-ZIP-CONTAINER")
    manifest = write_valid_manifest(base, workbook)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is False
    assert "INVALID_XLSX_CONTAINER" in finding_codes(report)


def test_missing_required_part_fails(tmp_path: Path) -> None:
    base = tmp_path / "base"
    base.mkdir()
    workbook = base / WORKBOOK_NAME
    with zipfile.ZipFile(workbook, "w") as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
    manifest = write_valid_manifest(base, workbook)
    assert "XLSX_REQUIRED_PART_MISSING" in finding_codes(run_inspect(workbook, manifest))


def test_traversal_entry_fails(tmp_path: Path) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    tampered = base / "files" / "TEST-TAMPERED.xlsx"
    add_zip_entry(workbook, tampered, "../TEST-ESCAPE.txt", b"TEST")
    manifest = write_valid_manifest(base, tampered)
    assert "XLSX_PATH_TRAVERSAL_ENTRY" in finding_codes(run_inspect(tampered, manifest))


def test_duplicate_entry_fails(tmp_path: Path) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    tampered = base / "files" / "TEST-TAMPERED.xlsx"
    with zipfile.ZipFile(workbook) as zin, zipfile.ZipFile(tampered, "w") as zout:
        for info in zin.infolist():
            zout.writestr(info, zin.read(info.filename))
        zout.writestr("xl/TEST-DUP.xml", b"<a/>")
        zout.writestr("xl/TEST-DUP.xml", b"<b/>")
    manifest = write_valid_manifest(base, tampered)
    assert "XLSX_DUPLICATE_ENTRY" in finding_codes(run_inspect(tampered, manifest))


def _tampered_fixture(tmp_path: Path, entry: str, data: bytes) -> tuple[Path, Path]:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    tampered = base / "files" / "TEST-TAMPERED.xlsx"
    add_zip_entry(workbook, tampered, entry, data)
    manifest = write_valid_manifest(base, tampered)
    return tampered, manifest


def test_macro_entry_detected(tmp_path: Path) -> None:
    workbook, manifest = _tampered_fixture(tmp_path, "xl/vbaProject.bin", b"TEST-MACRO")
    report = run_inspect(workbook, manifest)
    assert "WORKBOOK_MACRO_DETECTED" in finding_codes(report)
    assert report["summary"]["inspectionCompleted"] is False


def test_external_link_detected_and_target_hidden(tmp_path: Path) -> None:
    workbook, manifest = _tampered_fixture(
        tmp_path,
        "xl/externalLinks/externalLink1.xml",
        b"<externalLink>urn:test:external-link-target</externalLink>",
    )
    report = run_inspect(workbook, manifest)
    assert "WORKBOOK_EXTERNAL_LINK_DETECTED" in finding_codes(report)
    assert "urn:test:external-link-target" not in inspect_excel.serialize_report(report)


def test_embedded_object_detected(tmp_path: Path) -> None:
    workbook, manifest = _tampered_fixture(tmp_path, "xl/embeddings/oleObject1.bin", b"TEST-OLE")
    assert "WORKBOOK_EMBEDDED_OBJECT_DETECTED" in finding_codes(run_inspect(workbook, manifest))


def test_activex_detected(tmp_path: Path) -> None:
    workbook, manifest = _tampered_fixture(tmp_path, "xl/activeX/activeX1.xml", b"<ax/>")
    assert "WORKBOOK_ACTIVEX_DETECTED" in finding_codes(run_inspect(workbook, manifest))


# --- E. Manifest Gate ---


def test_verified_manifest_inspection(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    report = run_inspect(workbook, manifest)
    assert report["provenanceVerified"] is True
    assert report["manifestStatus"] == "verified"
    assert report["inspectionMode"] == "verified-source"
    assert report["sourceId"] == "TEST-SOURCE-001"


def test_gate_blocks_inspection_on_hash_mismatch(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    manifest = write_mismatch_manifest(base, workbook)

    def _must_not_run(path: Path, options: object) -> tuple[dict[str, Any], list[Any], bool]:
        raise AssertionError("Manifest Gate 실패 시 ZIP Preflight를 호출하면 안 된다")

    monkeypatch.setattr(inspect_excel, "_inspect_zip_container", _must_not_run)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is False
    assert report["manifestStatus"] == "invalid"


def test_gate_blocks_on_duplicate_manifest_key(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    manifest = base / "source-manifest.yaml"
    manifest.write_text("version: 1\nversion: 2\nsources: []\n", encoding="utf-8")

    def _must_not_run(path: Path, options: object) -> tuple[dict[str, Any], list[Any], bool]:
        raise AssertionError("Manifest Gate 실패 시 ZIP Preflight를 호출하면 안 된다")

    monkeypatch.setattr(inspect_excel, "_inspect_zip_container", _must_not_run)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is False
    assert report["manifestStatus"] == "invalid"


def test_pending_manifest_rejected_by_default(tmp_path: Path) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    manifest = write_pending_manifest(base)
    report = run_inspect(workbook, manifest)
    assert report["summary"]["inspectionCompleted"] is False
    assert report["manifestStatus"] == "pending"


def test_pending_manifest_allowed_with_override(tmp_path: Path) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    manifest = write_pending_manifest(base)
    report = run_inspect(workbook, manifest, allow_pending_manifest=True)
    assert report["summary"]["inspectionCompleted"] is True
    assert report["provenanceVerified"] is False
    assert report["manifestStatus"] == "pending"
    assert report["inspectionMode"] == "local-unverified-source"
    assert report["sourceId"] is None
    assert report["summary"]["normalizationReady"] is False
    assert report["summary"]["humanReviewRequired"] is True


def test_manifest_error_values_not_exposed(tmp_path: Path) -> None:
    base = tmp_path / "base"
    (base / "files").mkdir(parents=True)
    workbook = base / "files" / WORKBOOK_NAME
    build_normal_workbook(workbook)
    manifest = write_mismatch_manifest(base, workbook)
    serialized = inspect_excel.serialize_report(run_inspect(workbook, manifest))
    assert "f" * 64 not in serialized
    assert sha256_of_file(workbook) not in serialized


# --- F. Scan Limit 및 Output ---


def test_scan_limits_applied(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    sheet = run_inspect(workbook, manifest, max_rows_per_sheet=2, max_columns_per_sheet=2)[
        "sheets"
    ][0]
    assert sheet["scannedRowCount"] == 2
    assert sheet["scannedColumnCount"] == 2
    assert sheet["maxRowDeclared"] >= 3
    assert sheet["maxColumnDeclared"] >= 4


def test_scan_limit_finding(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    report = run_inspect(workbook, manifest, max_rows_per_sheet=2, max_columns_per_sheet=2)
    assert "WORKBOOK_SCAN_LIMIT_REACHED" in finding_codes(report)


def test_zero_max_rows_fails(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    report = run_inspect(workbook, manifest, max_rows_per_sheet=0)
    assert report["summary"]["inspectionCompleted"] is False
    assert "INVALID_INSPECTION_OPTION" in finding_codes(report)


def test_negative_max_columns_fails(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    report = run_inspect(workbook, manifest, max_columns_per_sheet=-1)
    assert report["summary"]["inspectionCompleted"] is False
    assert "INVALID_INSPECTION_OPTION" in finding_codes(report)


def test_cli_output_write_succeeds(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    output = tmp_path / "report.local.json"
    argv = [
        str(workbook), "--manifest", str(manifest),
        "--max-rows-per-sheet", "50", "--max-columns-per-sheet", "20",
        "--output", str(output),
    ]
    assert inspect_excel.main(argv) == 0
    text = output.read_text(encoding="utf-8")
    assert text.endswith("\n")
    assert WORKBOOK_NAME not in text


def test_output_conflict_with_workbook_rejected(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    argv = [
        str(workbook), "--manifest", str(manifest),
        "--max-rows-per-sheet", "50", "--max-columns-per-sheet", "20",
        "--output", str(workbook),
    ]
    assert inspect_excel.main(argv) == 1
    assert sha256_of_file(workbook) == sha256_of_file(workbook)


def test_output_conflict_with_manifest_rejected(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    before = manifest.read_bytes()
    argv = [
        str(workbook), "--manifest", str(manifest),
        "--max-rows-per-sheet", "50", "--max-columns-per-sheet", "20",
        "--output", str(manifest),
    ]
    assert inspect_excel.main(argv) == 1
    assert manifest.read_bytes() == before


def test_failure_keeps_workbook_unchanged(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    before = sha256_of_file(workbook)
    run_inspect(workbook, manifest, max_rows_per_sheet=0)
    assert sha256_of_file(workbook) == before


def test_failure_findings_contract(tmp_path: Path) -> None:
    workbook, manifest = make_fixture(tmp_path)
    failures = [
        run_inspect(workbook, manifest, max_rows_per_sheet=0),
        inspect_excel.inspect_workbook(
            "TEST-MARKER-PATH",  # type: ignore[arg-type]
            manifest_path=manifest,
            options=make_options(),
        ),
    ]
    for report in failures:
        assert report["summary"]["inspectionCompleted"] is False
        for finding in report["findings"]:
            assert finding["actualValue"] is None
        serialized = inspect_excel.serialize_report(report)
        assert "TEST-MARKER-PATH" not in serialized
