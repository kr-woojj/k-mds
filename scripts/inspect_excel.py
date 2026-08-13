"""IMO Compendium Workbook Inspector (ADR-0007).

read-only, deterministic, fail-closed Workbook Structure Inspector.

- source_manifest_load PASS가 기본 선행조건이다 (pending은 명시적 Override).
- ZIP Container Preflight를 openpyxl Load보다 먼저 수행한다.
- Cell 값, Formula 원문, Sheet 이름, Header 원문, 파일 이름, 경로, Hash를
  Report와 Console에 출력하지 않는다. Sheet 이름과 Header는 SHA-256 Digest로만
  표현한다.
- Formula 계산, Macro 실행, External Link Follow, Workbook 저장을 하지 않는다.
- 실제 원본 Report는 Internal Restricted Derived Metadata다.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
import yaml

from k_mds.models import ResultStatus
from k_mds.skills import source_manifest_load

# 중복 Key를 거부하는 Strict SafeLoader를 Loader 구현과 공유한다 (ADR-0006).
from k_mds.skills.source_manifest_load import _StrictSafeLoader

REPORT_VERSION = 1
_HEADER_SCAN_DEPTH = 10
_SHEET_XML_READ_LIMIT = 100 * 1024 * 1024
_CELL_CATEGORIES = (
    "blank", "string", "number", "boolean", "date",
    "formula", "error", "inline_string", "unknown",
)


@dataclass(frozen=True)
class InspectionOptions:
    max_rows_per_sheet: int
    max_columns_per_sheet: int
    allow_pending_manifest: bool = False
    fail_on_formulas: bool = False
    fail_on_external_links: bool = True
    fail_on_macros: bool = True
    fail_on_embedded_objects: bool = True
    fail_on_unsupported_feature: bool = False


def _build_finding(
    severity: str, code: str, message: str, path: str | None = None
) -> dict[str, Any]:
    return {
        "severity": severity,
        "code": code,
        "message": message,
        "path": path,
        "actualValue": None,
    }


def _digest_text(value: str) -> str:
    """Digest 계산 직전에만 정규화한다. 원문은 반환하지 않는다."""
    normalized = re.sub(r"\s+", " ", value.strip())
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _normalized_length(value: str) -> int:
    return len(re.sub(r"\s+", " ", value.strip()))


def _base_report() -> dict[str, Any]:
    return {
        "reportVersion": REPORT_VERSION,
        "deterministic": True,
        "fileRole": "imo-compendium-workbook",
        "provenanceVerified": False,
        "manifestStatus": "invalid",
        "inspectionMode": "not-inspected",
        "sourceId": None,
        "zipContainer": {},
        "workbook": {},
        "sheets": [],
        "findings": [],
        "summary": {
            "inspectionCompleted": False,
            "normalizationReady": False,
            "humanReviewRequired": True,
        },
    }


def _error_report(
    findings: list[dict[str, Any]], manifest_status: str = "invalid"
) -> dict[str, Any]:
    report = _base_report()
    report["manifestStatus"] = manifest_status
    report["findings"] = findings
    return report


# --- Manifest Gate ---

_PENDING_REQUIRED_ROOT_KEYS = {"standard", "files", "ingestion"}
_PENDING_ALLOWED_ROOT_KEYS = {"standard", "source", "files", "ingestion"}


def _is_pending_placeholder(text: str) -> bool:
    """정확한 pending_source Placeholder Contract만 pending으로 인정한다.

    Comment나 임의 문자열의 pending_source Marker는 pending이 아니다.
    Duplicate Key와 Root Extra Field는 invalid로 처리된다.
    """
    try:
        raw = yaml.load(text, Loader=_StrictSafeLoader)  # noqa: S506 — SafeLoader 기반
    except yaml.YAMLError:
        return False
    if not isinstance(raw, dict):
        return False
    keys = set(raw)
    if not _PENDING_REQUIRED_ROOT_KEYS <= keys or not keys <= _PENDING_ALLOWED_ROOT_KEYS:
        return False
    if raw.get("files") != []:
        return False
    standard = raw.get("standard")
    ingestion = raw.get("ingestion")
    if not isinstance(standard, dict) or standard.get("status") != "pending_source":
        return False
    if not isinstance(ingestion, dict) or ingestion.get("status") != "pending_source":
        return False
    return True


def _detect_manifest_state(
    manifest_path: Path, workbook_path: Path, base_dir: Path | None
) -> tuple[str, str | None, list[dict[str, Any]]]:
    """Manifest 상태를 verified / pending / invalid로 분류한다."""
    result = source_manifest_load(manifest_path, base_dir=base_dir)
    if result.status is ResultStatus.PASS:
        target = workbook_path.resolve()
        base = (base_dir if base_dir is not None else manifest_path.parent).resolve()
        for source in result.data["sources"]:
            if (base / str(source["source_file"])).resolve() == target:
                return "verified", str(source["source_id"]), []
        return (
            "invalid",
            None,
            [
                _build_finding(
                    "ERROR",
                    "WORKBOOK_NOT_IN_MANIFEST",
                    "Workbook이 Manifest의 검증된 Source Entry에 포함되지 않는다",
                    "$.manifest",
                )
            ],
        )

    try:
        text = manifest_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        text = ""
    if text and _is_pending_placeholder(text):
        return "pending", None, []

    findings = [
        _build_finding("ERROR", finding.code, finding.message, finding.path)
        for finding in result.errors
    ]
    return "invalid", None, findings


# --- ZIP Container Preflight ---


def _count_prefix(names: list[str], prefix: str) -> int:
    return sum(1 for name in names if name.startswith(prefix))


def _inspect_zip_container(
    path: Path, options: InspectionOptions
) -> tuple[dict[str, Any], list[dict[str, Any]], bool]:
    """openpyxl Load 전에 읽기 전용 ZIP Preflight를 수행한다.

    Entry 이름은 Report나 Console에 출력하지 않고 Category와 Count만 기록한다.
    """
    findings: list[dict[str, Any]] = []
    try:
        archive = zipfile.ZipFile(path)
    except (zipfile.BadZipFile, OSError):
        return (
            {},
            [
                _build_finding(
                    "ERROR",
                    "INVALID_XLSX_CONTAINER",
                    "Workbook이 유효한 XLSX(ZIP) Container가 아니다",
                    "$.zipContainer",
                )
            ],
            True,
        )

    with archive:
        infos = archive.infolist()
        names = [info.filename for info in infos]

        total_compressed = sum(info.compress_size for info in infos)
        total_uncompressed = sum(info.file_size for info in infos)
        max_entry_uncompressed = max((info.file_size for info in infos), default=0)
        encrypted_count = sum(1 for info in infos if info.flag_bits & 0x1)

        traversal = any(
            name.startswith("/")
            or re.match(r"^[A-Za-z]:", name)
            or ".." in name.replace("\\", "/").split("/")
            for name in names
        )
        duplicate = len(names) != len(set(names))

        content_types = ""
        if "[Content_Types].xml" in names:
            content_types = archive.read("[Content_Types].xml").decode("utf-8", "ignore")
        workbook_rels = ""
        if "xl/_rels/workbook.xml.rels" in names:
            workbook_rels = archive.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")

        macro_detected = (
            any(name.endswith("vbaProject.bin") for name in names)
            or "macroEnabled" in content_types
            or "vbaProject" in workbook_rels
        )
        external_link_part_count = _count_prefix(names, "xl/externalLinks/")
        external_link_rel_count = len(
            re.findall(r'Type="[^"]*/externalLink"', workbook_rels)
        )
        external_link_detected = bool(external_link_part_count or external_link_rel_count)
        embedded_count = _count_prefix(names, "xl/embeddings/") + sum(
            1 for name in names if "oleObject" in name
        )
        activex_count = _count_prefix(names, "xl/activeX/")
        custom_xml_count = _count_prefix(names, "customXml/")
        signature_count = _count_prefix(names, "_xmlsignatures/")
        drawing_part_count = sum(
            1
            for name in names
            if name.startswith("xl/drawings/") and "_rels" not in name
        )
        comment_part_count = sum(
            1
            for name in names
            if re.fullmatch(r"xl/comments(?:/comment)?\d*\.xml", name)
        )

    metadata: dict[str, Any] = {
        "entryCount": len(names),
        "totalCompressedSize": total_compressed,
        "totalUncompressedSize": total_uncompressed,
        "maxEntryUncompressedSize": max_entry_uncompressed,
        "compressionRatio": round(total_uncompressed / max(total_compressed, 1), 4),
        "encryptedEntryCount": encrypted_count,
        "macroDetected": macro_detected,
        "externalLinkDetected": external_link_detected,
        "externalLinkPartCount": external_link_part_count,
        "externalLinkRelationshipCount": external_link_rel_count,
        "embeddedObjectPartCount": embedded_count,
        "activeXPartCount": activex_count,
        "customXmlPartCount": custom_xml_count,
        "digitalSignaturePartCount": signature_count,
        "drawingPartCount": drawing_part_count,
        "commentPartCount": comment_part_count,
    }

    fatal = False
    if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
        findings.append(
            _build_finding(
                "ERROR",
                "XLSX_REQUIRED_PART_MISSING",
                "필수 XLSX Part가 누락되었다",
                "$.zipContainer",
            )
        )
        fatal = True
    if traversal:
        findings.append(
            _build_finding(
                "ERROR",
                "XLSX_PATH_TRAVERSAL_ENTRY",
                "ZIP Entry에 경로 이탈 이름이 존재한다",
                "$.zipContainer",
            )
        )
        fatal = True
    if duplicate:
        findings.append(
            _build_finding(
                "ERROR",
                "XLSX_DUPLICATE_ENTRY",
                "중복 ZIP Entry 이름이 존재한다",
                "$.zipContainer",
            )
        )
        fatal = True
    if encrypted_count:
        findings.append(
            _build_finding(
                "ERROR",
                "XLSX_ENCRYPTED_ENTRY",
                "암호화된 ZIP Entry가 존재한다",
                "$.zipContainer",
            )
        )
        fatal = True
    if macro_detected:
        severity = "ERROR" if options.fail_on_macros else "WARNING"
        findings.append(
            _build_finding(
                severity, "WORKBOOK_MACRO_DETECTED", "Macro Content가 탐지되었다", "$.zipContainer"
            )
        )
        fatal = fatal or options.fail_on_macros
    if external_link_detected:
        severity = "ERROR" if options.fail_on_external_links else "WARNING"
        findings.append(
            _build_finding(
                severity,
                "WORKBOOK_EXTERNAL_LINK_DETECTED",
                "External Link Part가 탐지되었다",
                "$.zipContainer",
            )
        )
        fatal = fatal or options.fail_on_external_links
    if embedded_count:
        severity = "ERROR" if options.fail_on_embedded_objects else "WARNING"
        findings.append(
            _build_finding(
                severity,
                "WORKBOOK_EMBEDDED_OBJECT_DETECTED",
                "Embedded 또는 OLE Object가 탐지되었다",
                "$.zipContainer",
            )
        )
        fatal = fatal or options.fail_on_embedded_objects
    if activex_count:
        severity = "ERROR" if options.fail_on_embedded_objects else "WARNING"
        findings.append(
            _build_finding(
                severity,
                "WORKBOOK_ACTIVEX_DETECTED",
                "ActiveX Content가 탐지되었다",
                "$.zipContainer",
            )
        )
        fatal = fatal or options.fail_on_embedded_objects
    if custom_xml_count:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_CUSTOM_XML_PRESENT",
                "Custom XML Part가 존재한다",
                "$.zipContainer",
            )
        )
    if signature_count:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_DIGITAL_SIGNATURE_PRESENT",
                "Digital Signature Part가 존재한다",
                "$.zipContainer",
            )
        )

    return metadata, findings, fatal


# --- Sheet 구조 보조 Metadata (ZIP XML Token Count) ---


def _read_zip_text(archive: zipfile.ZipFile, name: str) -> str | None:
    """Part 내용을 읽는다. 부재는 ""(정상), Read Limit 초과는 None(Skip)으로 구분한다."""
    try:
        info = archive.getinfo(name)
    except KeyError:
        return ""
    if info.file_size > _SHEET_XML_READ_LIMIT:
        return None
    return archive.read(name).decode("utf-8", "ignore")


def _collect_structure(path: Path) -> dict[str, Any]:
    """workbook.xml과 Sheet XML에서 구조 Count만 추출한다 (값·이름 미저장).

    Read Limit를 초과한 Part는 빈 값으로 위장하지 않고 skippedPartCount로
    집계한다 (Silent Undercount 방지).
    """
    skipped_parts = 0

    def _read_or_skip(archive: zipfile.ZipFile, name: str) -> str:
        nonlocal skipped_parts
        text = _read_zip_text(archive, name)
        if text is None:
            skipped_parts += 1
            return ""
        return text

    with zipfile.ZipFile(path) as archive:
        workbook_xml = _read_or_skip(archive, "xl/workbook.xml")
        rels_xml = _read_or_skip(archive, "xl/_rels/workbook.xml.rels")
        core_xml = _read_or_skip(archive, "docProps/core.xml")
        app_xml = _read_or_skip(archive, "docProps/app.xml")

        rel_targets: dict[str, str] = {}
        for rel_tag_match in re.finditer(r"<Relationship\b[^>]*>", rels_xml):
            rel_tag = rel_tag_match.group(0)
            rid_match = re.search(r'Id="([^"]+)"', rel_tag)
            target_match = re.search(r'Target="([^"]+)"', rel_tag)
            if rid_match and target_match:
                rel_targets[rid_match.group(1)] = target_match.group(1)

        sheet_states: list[str] = []
        sheet_parts: list[str] = []
        for sheet_tag_match in re.finditer(r"<sheet\b[^>]*/?>", workbook_xml):
            sheet_tag = sheet_tag_match.group(0)
            state_match = re.search(r'state="([^"]+)"', sheet_tag)
            sheet_states.append(state_match.group(1) if state_match else "visible")
            sheet_rid_match = re.search(r'r:id="([^"]+)"', sheet_tag)
            part = (
                rel_targets.get(sheet_rid_match.group(1), "") if sheet_rid_match else ""
            )
            part = part.lstrip("/")
            if part and not part.startswith("xl/"):
                part = f"xl/{part}"
            sheet_parts.append(part)

        per_sheet: list[dict[str, int | bool]] = []
        for part in sheet_parts:
            sheet_xml = _read_or_skip(archive, part) if part else ""
            rels_name = ""
            if part:
                parent, _, base_name = part.rpartition("/")
                rels_name = f"{parent}/_rels/{base_name}.rels"
            sheet_rels = _read_or_skip(archive, rels_name) if rels_name else ""
            per_sheet.append(
                {
                    "mergedRangeCount": len(re.findall(r"<mergeCell\b", sheet_xml)),
                    "hiddenRowCount": len(
                        re.findall(r'<row\b[^>]*hidden="(?:1|true)"', sheet_xml)
                    ),
                    "hiddenColumnCount": len(
                        re.findall(r'<col\b[^>]*hidden="(?:1|true)"', sheet_xml)
                    ),
                    "hyperlinkCount": len(re.findall(r"<hyperlink\b", sheet_xml)),
                    "dataValidationCount": len(
                        re.findall(r"<dataValidation\b", sheet_xml)
                    ),
                    "tableCount": len(re.findall(r"<tablePart\b", sheet_xml)),
                    "drawingCount": len(re.findall(r"<drawing\b", sheet_xml)),
                    "commentCount": len(
                        re.findall(r'Type="[^"]*/comments"', sheet_rels)
                    ),
                    "hasSheetProtection": "<sheetProtection" in sheet_xml,
                    "mergedTopRows": len(
                        [
                            row
                            for row in re.findall(
                                r'<mergeCell ref="[A-Z]{1,3}(\d+)', sheet_xml
                            )
                            if int(row) <= _HEADER_SCAN_DEPTH
                        ]
                    ),
                }
            )

    return {
        "sheetStates": sheet_states,
        "perSheet": per_sheet,
        "skippedPartCount": skipped_parts,
        "definedNameCount": len(re.findall(r"<definedName\b", workbook_xml)),
        "hasWorkbookProtection": "<workbookProtection" in workbook_xml,
        "dateSystem": "1904" if 'date1904="1"' in workbook_xml else "1900",
        "calculationMode": (
            calc_match.group(1)
            if (calc_match := re.search(r'<calcPr\b[^>]*calcMode="([^"]+)"', workbook_xml))
            else None
        ),
        "documentPropertiesPresent": {
            "creator": bool(re.search(r"<dc:creator>[^<]+</dc:creator>", core_xml)),
            "lastModifiedBy": bool(
                re.search(r"<cp:lastModifiedBy>[^<]+</cp:lastModifiedBy>", core_xml)
            ),
            "title": bool(re.search(r"<dc:title>[^<]+</dc:title>", core_xml)),
            "company": bool(re.search(r"<Company>[^<]+</Company>", app_xml)),
        },
    }


# --- Cell 및 Header ---


def _classify_cell(cell: Any) -> str:
    value = cell.value
    if value is None:
        return "blank"
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        return "formula"
    if data_type == "e":
        return "error"
    if data_type == "b":
        return "boolean"
    if not isinstance(value, str) and getattr(cell, "is_date", False):
        return "date"
    if data_type == "n":
        return "number"
    if data_type == "inlineStr":
        return "inline_string"
    if data_type in ("s", "str"):
        return "string"
    return "unknown"


def _infer_header(
    top_rows: list[list[tuple[str, str | None]]],
    merged_top_rows: int,
    sheet_ordinal: int,
    total_scanned_rows: int,
) -> tuple[int | None, str, list[dict[str, Any]], int, list[dict[str, Any]]]:
    """구조 기반 휴리스틱으로 Header Row를 추정한다. 원문은 저장하지 않는다."""
    findings: list[dict[str, Any]] = []
    for row_index, cells in enumerate(top_rows, start=1):
        non_empty = [cell for cell in cells if cell[0] != "blank"]
        if len(non_empty) < 2:
            continue
        strings = [cell for cell in cells if cell[1] is not None]
        string_ratio = len(strings) / len(non_empty)
        if string_ratio < 0.5:
            continue

        digests = [_digest_text(text) for _, text in strings if text is not None]
        duplicate_count = len(digests) - len(set(digests))
        confidence = "high" if string_ratio >= 0.8 else "medium"
        if row_index >= total_scanned_rows:
            confidence = "medium"
        if merged_top_rows:
            findings.append(
                _build_finding(
                    "WARNING",
                    "WORKBOOK_MERGED_HEADER_REGION",
                    "Header 후보 영역에 병합 셀이 존재한다",
                    f"$.sheets.{sheet_ordinal}.rows.{row_index}",
                )
            )
            if confidence == "high":
                confidence = "medium"
        if duplicate_count:
            findings.append(
                _build_finding(
                    "WARNING",
                    "WORKBOOK_DUPLICATE_HEADER_DIGEST",
                    "Header 후보에 중복 Digest가 존재한다",
                    f"$.sheets.{sheet_ordinal}.rows.{row_index}",
                )
            )

        candidates = [
            {
                "row": row_index,
                "column": column_index + 1,
                "valueDigest": _digest_text(text),
                "normalizedLength": _normalized_length(text),
                "cellType": cell_type,
            }
            for column_index, (cell_type, text) in enumerate(cells)
            if text is not None
        ]
        return row_index, confidence, candidates, duplicate_count, findings

    findings.append(
        _build_finding(
            "WARNING",
            "WORKBOOK_HEADER_NOT_DETECTED",
            "구조 휴리스틱으로 Header Row를 찾지 못했다",
            f"$.sheets.{sheet_ordinal}",
        )
    )
    return None, "none", [], 0, findings


def _inspect_sheet(
    worksheet: Any,
    sheet_ordinal: int,
    state: str,
    structure: dict[str, int | bool],
    options: InspectionOptions,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    findings: list[dict[str, Any]] = []
    declared_rows = int(worksheet.max_row or 0)
    declared_columns = int(worksheet.max_column or 0)
    scan_rows = min(declared_rows, options.max_rows_per_sheet) if declared_rows else 0
    scan_columns = (
        min(declared_columns, options.max_columns_per_sheet) if declared_columns else 0
    )

    type_counts = dict.fromkeys(_CELL_CATEGORIES, 0)
    non_empty_total = 0
    formula_count = 0
    error_count = 0
    blank_rows = 0
    scanned_rows = 0
    column_non_empty = [0] * max(scan_columns, 1)
    top_rows: list[list[tuple[str, str | None]]] = []

    if scan_rows and scan_columns:
        for row in worksheet.iter_rows(
            min_row=1, max_row=scan_rows, max_col=scan_columns
        ):
            scanned_rows += 1
            row_non_empty = 0
            row_data: list[tuple[str, str | None]] = []
            for column_index, cell in enumerate(row):
                category = _classify_cell(cell)
                type_counts[category] += 1
                if category != "blank":
                    row_non_empty += 1
                    non_empty_total += 1
                    if column_index < len(column_non_empty):
                        column_non_empty[column_index] += 1
                    if category == "formula":
                        formula_count += 1
                    elif category == "error":
                        error_count += 1
                if scanned_rows <= _HEADER_SCAN_DEPTH:
                    text = (
                        str(cell.value)
                        if category in ("string", "inline_string") and cell.value is not None
                        else None
                    )
                    row_data.append((category, text))
            if scanned_rows <= _HEADER_SCAN_DEPTH:
                top_rows.append(row_data)
            if row_non_empty == 0:
                blank_rows += 1

    blank_columns = sum(
        1 for count in column_non_empty[:scan_columns] if count == 0
    ) if scan_columns else 0

    # Scan Limit은 실제 iterated row 수가 아니라 선언 Dimension이
    # 명시적 Scan Budget을 초과한 경우에만 발생한다 (ADR-0007 Amendment).
    scan_budget_truncated = (
        declared_rows > options.max_rows_per_sheet
        or declared_columns > options.max_columns_per_sheet
    )

    # Empty·Drawing-only 판정은 Full Scan Coverage가 확보된 경우에만 가능하다.
    record_free = (
        not scan_budget_truncated
        and non_empty_total == 0
        and formula_count == 0
        and error_count == 0
        and all(
            int(structure[key]) == 0
            for key in (
                "mergedRangeCount",
                "hyperlinkCount",
                "dataValidationCount",
                "tableCount",
                "commentCount",
            )
        )
    )
    drawing_count = int(structure["drawingCount"])
    is_empty_sheet = record_free and drawing_count == 0
    # Drawing-only는 Empty가 아니라 Non-tabular Sheet Candidate다 (ADR-0007 Amendment).
    is_drawing_only_sheet = record_free and drawing_count >= 1

    if is_empty_sheet or is_drawing_only_sheet:
        inferred_row = None
        confidence = "none"
        candidates: list[dict[str, Any]] = []
        duplicate_count = 0
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_EMPTY_SHEET",
                "Sheet에 정규화 가능한 Cell Record가 존재하지 않는다",
                f"$.sheets.{sheet_ordinal}",
            )
            if is_empty_sheet
            else _build_finding(
                "WARNING",
                "WORKBOOK_DRAWING_ONLY_SHEET",
                "Sheet에 정규화 가능한 Cell Record 없이 Drawing Content가 존재한다",
                f"$.sheets.{sheet_ordinal}",
            )
        )
    else:
        inferred_row, confidence, candidates, duplicate_count, header_findings = (
            _infer_header(
                top_rows, int(structure["mergedTopRows"]), sheet_ordinal, scanned_rows
            )
        )
        findings.extend(header_findings)

    if scan_budget_truncated:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_SCAN_LIMIT_REACHED",
                "선언된 Dimension이 Scan Limit를 초과하여 일부만 검사했다",
                f"$.sheets.{sheet_ordinal}",
            )
        )
    if (
        declared_rows > 10 * options.max_rows_per_sheet
        or declared_columns > 10 * options.max_columns_per_sheet
    ):
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_DECLARED_DIMENSION_EXCESSIVE",
                "선언된 Dimension이 Scan Limit 대비 과도하다",
                f"$.sheets.{sheet_ordinal}",
            )
        )
    if error_count:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_CELL_ERROR_PRESENT",
                "Error Cell이 존재한다",
                f"$.sheets.{sheet_ordinal}",
            )
        )

    sheet_report = {
        "sheetOrdinal": sheet_ordinal,
        "sheetNameDigest": hashlib.sha256(
            str(worksheet.title).encode("utf-8")
        ).hexdigest(),
        "state": state,
        "maxRowDeclared": declared_rows,
        "maxColumnDeclared": declared_columns,
        "scannedRowCount": scanned_rows,
        "scannedColumnCount": scan_columns,
        "nonEmptyCellCount": non_empty_total,
        "formulaCellCount": formula_count,
        "errorCellCount": error_count,
        "mergedRangeCount": structure["mergedRangeCount"],
        "hiddenRowCount": structure["hiddenRowCount"],
        "hiddenColumnCount": structure["hiddenColumnCount"],
        "commentCount": structure["commentCount"],
        "hyperlinkCount": structure["hyperlinkCount"],
        "dataValidationCount": structure["dataValidationCount"],
        "tableCount": structure["tableCount"],
        "drawingCount": structure["drawingCount"],
        "blankRowCountInScan": blank_rows,
        "blankColumnCountInScan": blank_columns,
        "duplicateHeaderDigestCount": duplicate_count,
        "inferredHeaderRow": inferred_row,
        "headerConfidence": confidence,
        "headerCandidates": candidates,
        "cellTypeCounts": type_counts,
        "findings": [],
    }
    return sheet_report, findings


# --- Public API ---


def inspect_workbook(
    path: Path,
    *,
    manifest_path: Path,
    source_base_dir: Path | None = None,
    options: InspectionOptions,
) -> dict[str, Any]:
    """Workbook을 read-only로 검사하고 결정론적 Report Dict를 반환한다.

    source_base_dir가 None이면 manifest_path.parent를 Source Base로 사용한다.
    지정된 경우 Loader의 base_dir로 전달되며 Report에는 포함하지 않는다.
    """
    if not isinstance(path, Path):
        return _error_report(
            [_build_finding("ERROR", "WORKBOOK_PATH_NOT_PATH", "workbook path는 Path여야 한다")]
        )
    if not isinstance(manifest_path, Path):
        return _error_report(
            [_build_finding("ERROR", "MANIFEST_PATH_NOT_PATH", "manifest path는 Path여야 한다")]
        )
    if source_base_dir is not None and not isinstance(source_base_dir, Path):
        return _error_report(
            [
                _build_finding(
                    "ERROR",
                    "SOURCE_BASE_DIR_NOT_PATH",
                    "source_base_dir는 Path여야 한다",
                    "$.sourceBaseDir",
                )
            ]
        )
    if options.max_rows_per_sheet <= 0 or options.max_columns_per_sheet <= 0:
        return _error_report(
            [
                _build_finding(
                    "ERROR",
                    "INVALID_INSPECTION_OPTION",
                    "Scan Limit는 양의 정수여야 한다",
                    "$.options",
                )
            ]
        )
    if not path.is_file():
        return _error_report(
            [_build_finding("ERROR", "WORKBOOK_FILE_NOT_FOUND", "Workbook 파일이 존재하지 않는다")]
        )
    if path.suffix.lower() != ".xlsx":
        return _error_report(
            [_build_finding("ERROR", "WORKBOOK_NOT_XLSX", "Workbook은 XLSX 파일이어야 한다")]
        )
    if not manifest_path.is_file():
        return _error_report(
            [_build_finding("ERROR", "MANIFEST_FILE_NOT_FOUND", "Manifest 파일이 존재하지 않는다")]
        )
    if source_base_dir is not None:
        if not source_base_dir.is_dir():
            return _error_report(
                [
                    _build_finding(
                        "ERROR",
                        "SOURCE_BASE_DIR_NOT_FOUND",
                        "source_base_dir가 존재하는 Directory가 아니다",
                        "$.sourceBaseDir",
                    )
                ]
            )
        if not path.resolve().is_relative_to(source_base_dir.resolve()):
            return _error_report(
                [
                    _build_finding(
                        "ERROR",
                        "WORKBOOK_OUTSIDE_SOURCE_BASE",
                        "Workbook이 source_base_dir 아래에 있지 않다",
                        "$.sourceBaseDir",
                    )
                ]
            )

    manifest_status, source_id, gate_findings = _detect_manifest_state(
        manifest_path, path, source_base_dir
    )
    if manifest_status == "invalid":
        return _error_report(gate_findings, manifest_status="invalid")
    if manifest_status == "pending" and not options.allow_pending_manifest:
        return _error_report(
            [
                _build_finding(
                    "ERROR",
                    "MANIFEST_NOT_VERIFIED",
                    "Manifest가 pending 상태다 (--allow-pending-manifest가 필요하다)",
                    "$.manifest",
                )
            ],
            manifest_status="pending",
        )

    report = _base_report()
    report["manifestStatus"] = manifest_status
    report["provenanceVerified"] = manifest_status == "verified"
    report["inspectionMode"] = (
        "verified-source" if manifest_status == "verified" else "local-unverified-source"
    )
    report["sourceId"] = source_id

    zip_metadata, zip_findings, zip_fatal = _inspect_zip_container(path, options)
    report["zipContainer"] = zip_metadata
    findings: list[dict[str, Any]] = list(zip_findings)
    if zip_fatal:
        report["findings"] = findings
        return report

    structure = _collect_structure(path)
    skipped_part_count = int(structure["skippedPartCount"])
    if skipped_part_count:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_XML_PART_SCAN_SKIPPED",
                "Read Limit를 초과한 XML Part의 구조 Count를 생략했다",
                "$.workbook",
            )
        )

    try:
        workbook = openpyxl.load_workbook(
            filename=path, read_only=True, data_only=False, keep_links=False
        )
    except Exception:  # noqa: BLE001 — 예외를 정규화한다
        findings.append(
            _build_finding(
                "ERROR", "WORKBOOK_LOAD_FAILED", "Workbook을 read-only로 열 수 없다"
            )
        )
        report["findings"] = findings
        return report

    sheets: list[dict[str, Any]] = []
    try:
        states: list[str] = structure["sheetStates"]
        per_sheet: list[dict[str, int | bool]] = structure["perSheet"]
        for ordinal, worksheet in enumerate(workbook.worksheets):
            state = states[ordinal] if ordinal < len(states) else "visible"
            sheet_structure = (
                per_sheet[ordinal]
                if ordinal < len(per_sheet)
                else {
                    "mergedRangeCount": 0,
                    "hiddenRowCount": 0,
                    "hiddenColumnCount": 0,
                    "hyperlinkCount": 0,
                    "dataValidationCount": 0,
                    "tableCount": 0,
                    "drawingCount": 0,
                    "commentCount": 0,
                    "hasSheetProtection": False,
                    "mergedTopRows": 0,
                }
            )
            sheet_report, sheet_findings = _inspect_sheet(
                worksheet, ordinal, state, sheet_structure, options
            )
            sheets.append(sheet_report)
            findings.extend(sheet_findings)
    finally:
        workbook.close()

    hidden_count = sum(1 for state in states if state == "hidden")
    very_hidden_count = sum(1 for state in states if state == "veryHidden")
    formula_total = sum(int(sheet["formulaCellCount"]) for sheet in sheets)
    has_sheet_protection = any(bool(item["hasSheetProtection"]) for item in per_sheet)

    if hidden_count:
        findings.append(
            _build_finding(
                "WARNING", "WORKBOOK_HIDDEN_SHEET_PRESENT", "Hidden Sheet가 존재한다", "$.workbook"
            )
        )
    if very_hidden_count:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_VERY_HIDDEN_SHEET_PRESENT",
                "VeryHidden Sheet가 존재한다 (Human Review 필요)",
                "$.workbook",
            )
        )
    if formula_total:
        findings.append(
            _build_finding(
                "WARNING", "WORKBOOK_FORMULA_PRESENT", "Formula Cell이 존재한다", "$.workbook"
            )
        )
    if has_sheet_protection or structure["hasWorkbookProtection"]:
        findings.append(
            _build_finding(
                "WARNING",
                "WORKBOOK_PROTECTION_ENABLED",
                "Workbook 또는 Sheet Protection이 활성화되어 있다",
                "$.workbook",
            )
        )

    report["workbook"] = {
        "sheetCount": len(sheets),
        "visibleSheetCount": len(sheets) - hidden_count - very_hidden_count,
        "hiddenSheetCount": hidden_count,
        "veryHiddenSheetCount": very_hidden_count,
        "definedNameCount": structure["definedNameCount"],
        "externalLinkDetected": zip_metadata.get("externalLinkDetected", False),
        "externalLinkPartCount": zip_metadata.get("externalLinkPartCount", 0),
        "externalLinkRelationshipCount": zip_metadata.get(
            "externalLinkRelationshipCount", 0
        ),
        "formulaCellCount": formula_total,
        "mergedRangeCount": sum(int(item["mergedRangeCount"]) for item in per_sheet),
        "hiddenRowCount": sum(int(item["hiddenRowCount"]) for item in per_sheet),
        "hiddenColumnCount": sum(int(item["hiddenColumnCount"]) for item in per_sheet),
        "commentCount": zip_metadata.get("commentPartCount", 0),
        "hyperlinkCount": sum(int(item["hyperlinkCount"]) for item in per_sheet),
        "dataValidationCount": sum(int(item["dataValidationCount"]) for item in per_sheet),
        "tableCount": sum(int(item["tableCount"]) for item in per_sheet),
        "drawingCount": zip_metadata.get("drawingPartCount", 0),
        "embeddedObjectCount": zip_metadata.get("embeddedObjectPartCount", 0),
        "activeXCount": zip_metadata.get("activeXPartCount", 0),
        "customXmlCount": zip_metadata.get("customXmlPartCount", 0),
        "digitalSignatureCount": zip_metadata.get("digitalSignaturePartCount", 0),
        "macroDetected": zip_metadata.get("macroDetected", False),
        "hasWorkbookProtection": structure["hasWorkbookProtection"],
        "hasSheetProtection": has_sheet_protection,
        "unsupportedFeatureCount": skipped_part_count,
        "documentPropertiesPresent": structure["documentPropertiesPresent"],
        "dateSystem": structure["dateSystem"],
        "calculationMode": structure["calculationMode"],
    }
    report["sheets"] = sheets
    report["findings"] = findings

    has_error = any(finding["severity"] == "ERROR" for finding in findings)
    report["summary"] = {
        "inspectionCompleted": True,
        "normalizationReady": (
            report["provenanceVerified"] and not has_error and not findings
        ),
        "humanReviewRequired": (not report["provenanceVerified"]) or bool(findings),
    }
    return report


def serialize_report(report: dict[str, object]) -> str:
    """결정론적 직렬화: UTF-8, indent=2, sort_keys, 마지막 newline."""
    return json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n"


def _validate_output_path(output: Path, workbook: Path, manifest: Path) -> str | None:
    resolved = output.resolve()
    if resolved == workbook.resolve() or resolved == manifest.resolve():
        return "OUTPUT_PATH_CONFLICT"
    if resolved.is_dir():
        return "OUTPUT_PATH_CONFLICT"
    if not resolved.parent.is_dir():
        return "OUTPUT_WRITE_FAILED"
    return None


def write_report(path: Path, report: dict[str, object]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(serialize_report(report))


def _print_safe_summary(report: dict[str, Any]) -> None:
    """이름·값·경로·Hash를 출력하지 않는 최소 Console 요약."""
    summary = report.get("summary", {})
    workbook = report.get("workbook", {})
    print(
        "[inspect] "
        f"completed={summary.get('inspectionCompleted')} "
        f"sheets={workbook.get('sheetCount', 0)} "
        f"provenanceVerified={report.get('provenanceVerified')} "
        f"manifestStatus={report.get('manifestStatus')}"
    )
    counts: dict[str, int] = {}
    for finding in report.get("findings", []):
        counts[str(finding.get("code"))] = counts.get(str(finding.get("code")), 0) + 1
    for code in sorted(counts):
        print(f"[inspect] finding {code}={counts[code]}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="inspect_excel.py",
        description="read-only IMO Compendium Workbook Structure Inspector (ADR-0007)",
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--source-base-dir", type=Path, default=None)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--allow-pending-manifest", action="store_true", default=False)
    parser.add_argument("--max-rows-per-sheet", type=int, required=True)
    parser.add_argument("--max-columns-per-sheet", type=int, required=True)
    parser.add_argument("--fail-on-formulas", action="store_true", default=False)
    parser.add_argument(
        "--fail-on-external-links", action=argparse.BooleanOptionalAction, default=True
    )
    parser.add_argument(
        "--fail-on-macros", action=argparse.BooleanOptionalAction, default=True
    )
    parser.add_argument(
        "--fail-on-embedded-objects", action=argparse.BooleanOptionalAction, default=True
    )
    parser.add_argument("--fail-on-unsupported-feature", action="store_true", default=False)
    args = parser.parse_args(argv)

    options = InspectionOptions(
        max_rows_per_sheet=args.max_rows_per_sheet,
        max_columns_per_sheet=args.max_columns_per_sheet,
        allow_pending_manifest=args.allow_pending_manifest,
        fail_on_formulas=args.fail_on_formulas,
        fail_on_external_links=args.fail_on_external_links,
        fail_on_macros=args.fail_on_macros,
        fail_on_embedded_objects=args.fail_on_embedded_objects,
        fail_on_unsupported_feature=args.fail_on_unsupported_feature,
    )

    output_error = _validate_output_path(args.output, args.workbook, args.manifest)
    if output_error == "OUTPUT_PATH_CONFLICT":
        print("[inspect] output path가 입력 파일과 충돌한다 (OUTPUT_PATH_CONFLICT)")
        return 1
    if output_error is not None:
        print("[inspect] output path에 쓸 수 없다 (OUTPUT_WRITE_FAILED)")
        return 1

    report = inspect_workbook(
        args.workbook,
        manifest_path=args.manifest,
        source_base_dir=args.source_base_dir,
        options=options,
    )

    try:
        write_report(args.output, report)
    except OSError:
        print("[inspect] Report 저장에 실패했다 (OUTPUT_WRITE_FAILED)")
        return 1

    _print_safe_summary(report)

    summary = report.get("summary", {})
    workbook_info: dict[str, Any] = (
        report.get("workbook", {}) if isinstance(report.get("workbook"), dict) else {}
    )
    if not summary.get("inspectionCompleted"):
        return 1
    if any(finding["severity"] == "ERROR" for finding in report.get("findings", [])):
        return 1
    if options.fail_on_formulas and int(workbook_info.get("formulaCellCount", 0)) > 0:
        return 1
    if (
        options.fail_on_unsupported_feature
        and int(workbook_info.get("unsupportedFeatureCount", 0)) > 0
    ):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
